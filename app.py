from flask import Flask, render_template, request, redirect, send_file
import sqlite3
import pandas as pd
import os
from datetime import datetime, date
from openpyxl import Workbook

app = Flask(__name__)
DB_PATH = "data/ladungen.db"

def init_db():
    os.makedirs("data", exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS ladezyklen (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            datum TEXT,
            kilometerstand INTEGER,
            ladekosten REAL,
            lademenge_kw REAL
        )
    ''')
    conn.commit()
    conn.close()

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        lademenge = float(request.form["lademenge_kw"])
        ladekosten = round(lademenge * 0.2756, 2)  # in Euro

        data = (
            request.form["datum"],
            request.form["kilometerstand"],
            ladekosten,
            lademenge
        )
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('''
            INSERT INTO ladezyklen
            (datum, kilometerstand, ladekosten, lademenge_kw)
            VALUES (?, ?, ?, ?)
        ''', data)
        conn.commit()
        conn.close()
        return redirect("/")

    conn = sqlite3.connect(DB_PATH)
    ladezyklen = conn.execute("SELECT * FROM ladezyklen ORDER BY datum DESC").fetchall()
    conn.close()
    return render_template("index.html", ladezyklen=ladezyklen, today=date.today().isoformat())

@app.route("/export", methods=["GET", "POST"])
def export():
    if request.method == "POST":
        monat = request.form["monat"]  # Format: YYYY-MM
        start = f"{monat}-01"
        end = f"{monat}-31"
        monat_text = datetime.strptime(start, "%Y-%m-%d").strftime("%B %Y")

        conn = sqlite3.connect(DB_PATH)
        df = pd.read_sql_query(
            "SELECT datum, kilometerstand, ladekosten, lademenge_kw FROM ladezyklen WHERE datum BETWEEN ? AND ?",
            conn,
            params=(start, end)
        )
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Ladebericht"

        # Kopfzeile (B2–F2)
        ws.merge_cells("B2:D2")
        ws["B2"] = "Abrechnung Stromkosten Firmenwagen Monat:"
        ws["E2"] = monat_text

        # Gesamtkosten
        ws["D4"] = "Gesamtkosten"
        ws["E4"] = "=SUMME(C12:C30)"  # Optional: Formelbereich anpassen

        # Stammdaten A3–B8
        ws["A3"] = "Abrechnungsmonat"
        ws["B3"] = monat_text
        ws["A4"] = "Kennzeichen"
        ws["B4"] = "EDIT"
        ws["A5"] = "Fahrer"
        ws["B5"] = "EDIT"
        ws["A6"] = "Abteilung"
        ws["B6"] = "EDIT"
        ws["A7"] = "Kostenstelle"
        ws["B7"] = "EDIT"
        ws["A8"] = "Kosten pro KW"
        ws["B8"] = "EDIT"

        # Unterschrift-Feld
        ws.merge_cells("C8:G8")
        ws["C8"] = "______________________________"
        ws["C9"] = "Unterschrift"
        ws["D9"] = "Datum"

        # Datenüberschriften (ab A11)
        headers = ["Datum", "Kilometerstand", "Ladekosten (€)", "Lademenge (kWh)"]
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=11, column=col_num, value=header)

        for row_num, row in enumerate(df.values, start=12):
            for col_num, cell_value in enumerate(row, start=1):
                ws.cell(row=row_num, column=col_num, value=cell_value)

        export_path = f"data/export_{monat}.xlsx"
        wb.save(export_path)

        return send_file(export_path, as_attachment=True)

    return render_template("export.html")
	
if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=5000)
